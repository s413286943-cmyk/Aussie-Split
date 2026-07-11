import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def write_workbook(input_path, workbook_path):
    payload = json.loads(Path(input_path).read_text(encoding="utf-8"))
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for sheet_name in ("Days", "Blocks", "Resources"):
        rows = payload[sheet_name]
        sheet = workbook.create_sheet(sheet_name)
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])

    Path(workbook_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)


def read_workbook(workbook_path):
    workbook = load_workbook(workbook_path, data_only=True)
    result = {}
    for sheet_name in ("Days", "Blocks", "Resources"):
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in rows[0]]
        result[sheet_name] = [
            {
                header: "" if value is None else value
                for header, value in zip(headers, row)
            }
            for row in rows[1:]
            if any(value is not None and value != "" for value in row)
        ]
    print(json.dumps(result, ensure_ascii=False))


def read_finance(workbook_path):
    workbook = load_workbook(workbook_path, data_only=True)
    snapshot = workbook["LedgerSnapshot"]
    exported_at = snapshot["B2"].value
    if hasattr(exported_at, "isoformat"):
        exported_at = exported_at.isoformat(timespec="milliseconds") + "Z"

    result = {
        "snapshot": {
            "exportedAt": exported_at,
            "activityCount": snapshot["J2"].value,
            "totalsByCurrency": {
                "AUD": {
                    "confirmed": snapshot["C6"].value,
                    "pendingSettlement": snapshot["C7"].value,
                    "splitSettled": snapshot["C8"].value,
                },
                "CNY": {
                    "confirmed": snapshot["B6"].value,
                    "pendingSettlement": snapshot["B7"].value,
                    "splitSettled": snapshot["B8"].value,
                },
            },
        },
        "lodging": [
            {"name": row[2], "price": row[4]}
            for row in workbook["Lodging"].iter_rows(min_row=2, max_col=8, values_only=True)
            if row[2]
        ],
        "activityCosts": [
            {"item": row[1], "price": row[3], "cny": row[4]}
            for row in workbook["ActivityCosts"].iter_rows(min_row=2, max_col=6, values_only=True)
            if row[1]
        ],
    }
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    command = sys.argv[1]
    if command == "write":
        write_workbook(sys.argv[2], sys.argv[3])
    elif command == "read":
        read_workbook(sys.argv[2])
    elif command == "read_finance":
        read_finance(sys.argv[2])
    else:
        raise SystemExit(f"Unknown command: {command}")
